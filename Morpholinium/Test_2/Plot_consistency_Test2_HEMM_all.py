# Plot_consistency - HEMM
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.font_manager import FontProperties
import pandas as pd
import numpy as np
import openpyxl
import os

matplotlib.rcParams['font.family'] = "Arial" # change the default font # mudança da fonte
matplotlib.rcParams['xtick.direction'] = 'in' # change the ticks direction # mudança da direção dos marcadores ao longo dos eixos de um gráfico que indicam valores específicos
matplotlib.rcParams['ytick.direction'] = 'in'
font = FontProperties() # cria uma instância de propriedades de fontes, tamanho, estilo e família
font.set_family('sans-serif') # 'serif', 'sans-serif', 'cursive', 'fantasy', or 'monospace' # família da fonte
font.set_name('Arial') # nome da fonte
font.set_weight('bold') # 'ultralight', 'light', 'normal', 'regular', 'book', 'medium', 'roman', 'semibold', 'demibold', 'demi', 'bold', 'heavy', 'extra bold', 'black' # peso da fonte
font.set_style('normal') #'normal', 'italic' or 'oblique' # estilo da fonte, negrito, itálico, normal, etc
font.set_size('x-large') # xx-small', 'x-small', 'small', 'medium', 'large', 'x-large', 'xx-large' # tamanho da fonte
plt.rcParams['figure.dpi'] = 300 # resolução padrão para as figuras exibidas na tela
plt.rcParams['savefig.dpi'] = 300 # resolução padrão para as figuras salvas em arquivos PNG, PDF ou SVG
opt = {'fillstyle':'full','markersize':5,'markeredgewidth':0.5} # aparência dos marcadores, fillstyle: full - estilo de preenchimento markersize: 5 - tamanho do marcador markeredgewidth: 0.5 largura da borda do marcador
plt.rcParams['figure.facecolor'] = (1.0,1.0,1.0,0.0) # cor de fundo da figura como transparente
plt.rcParams['axes.facecolor'] = '0.95' # cor de fundo dos eixos como uma tonalidade de cinza claro
#plt.rcParams['figure.figsize'] = (4,4) # tamanho da figura padrão em polegadas


caminho_arquivo = 'HEMM.xlsx' # caminho do arquivo
wb = openpyxl.load_workbook(caminho_arquivo) # ler dados de abas individuais do arquivo Excel sem abrir e fechar repetidamente, mantém aberto. Mais eficiente se você precisa ler dados de várias abas do arquivo. Lista com os nomes de todas as abas do arquivo Excel.
# load_workbook é a função usada para carregar o arquvio Excel

dados_coluna_O_array = [] # matriz para guardar os dados da coluna O de todas as sheet
dados_coluna_P_array = [] # matriz para guardar os dados da coluna P de todas as sheet
dados_coluna_Q_array = [] # matriz para guardar os dados da coluna Q de todas as sheet

nomes_abas = []

for i in wb.sheetnames: # para i dentro da lista wb com os nomes de todas as abas
    nome_aba = wb[i] # obtém o nome da aba na posição i na lista wb, que contém os nomes de todas as abas
    nome_aba_split = nome_aba.title.split("_") # nome_aba é o nome completo da aba. nome_aba.title converte o nome da aba para o formato de título. split("_ ") divide o nome da aba em duas partes usando o caractere "_ "
    if len(nome_aba_split) > 2: # se o nome da aba contiver mais de duas partes
        nome_aba_split.insert(2, "wt%") # insere wt% entre a segunda e a terceira partes do nome da aba
        nome_aba_final = "_".join(nome_aba_split) # junta as partes do nome da aba de volta com "_"
        print('nome_aba_final', nome_aba_final)
        nomes_abas.append(nome_aba_final) # adiciona o nome modificado _a lista de nomes das abas
    else:
        nomes_abas.append(nome_aba.title) # nome da aba não tiver mais de duas partes, adiciona o nome completo da aba à lista nomes_abas (usada na legenda)

    dados_coluna_O = [] # abre uma lista para os dados da coluna O da sheet
    dados_coluna_P = [] # abre uma lista para os dados da coluna P da sheet
    dados_coluna_Q = [] # abre uma lista para os dados da coluna Q da sheet

    for j in nome_aba.iter_rows(min_row=2, values_only=True): # iter_rows: permite iterar sobre as linhas dessa aba # min_row=2: especifica a primeira linha a ser considerada, nesse caso, começa na segunda linha, pois a primeira linha geralmente contém os cabeçalhos das colunas # values_only=True: indica que queremos obter apenas os valores das células, sem incluir informações adicionais como formatação, fórmulas ou estilos.
        dados_coluna_O.append(j[14])
        dados_coluna_P.append(j[15])
        dados_coluna_Q.append(j[16])

    dados_coluna_O_array.append(dados_coluna_O)
    dados_coluna_P_array.append(dados_coluna_P)
    dados_coluna_Q_array.append(dados_coluna_Q)

print('dados_O', dados_coluna_O_array)
print('dados_P', dados_coluna_P_array)
print('dados_Q', dados_coluna_Q_array)

# Plotagem dos dados:

symbols = ['o', 's', '^', 'D', 'v', '>', '<', 'P', 'X', 'H', '*', 'h', '+', 'x', '|', '_', '1', '2', '3', '4', 'p', ',', '.', '8', 'p', 'd', 'D', '|', '_', '$...$']
# Exemplos de símbolos

# Vetor de cores
colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',  '#aec7e8',  '#ffbb78', '#98df8a', '#ff9896', '#c5b0d5', '#c49c94',  '#f7b6d2', '#c7c7c7',  '#dbdb8d',  '#9edae5',  '#393b79',  '#5254a3', '#6b6ecf',  '#9c9ede',  '#637939',  '#8ca252',  '#b5cf6b', '#cedb9c',  '#8c6d31',  '#bd9e39',  '#e7ba52',  '#e7cb94',  '#843c39', '#ad494a', '#d6616b',  '#e7969c',  '#7b4173', '#a55194',  '#ce6dbd',  '#de9ed6']

lines = []

for i in range(len(dados_coluna_O_array)):
    y = dados_coluna_O_array[i]
    x = dados_coluna_P_array[i]
    x2 = dados_coluna_Q_array[i]
    print('x', x)
    print('y', y)
    print('x2', x2)
    symbol=symbols[i]
    color=colors[i]
    line1, = plt.plot(x, y, marker=symbol, markersize=opt['markersize'], markeredgewidth=opt['markeredgewidth'], fillstyle=opt['fillstyle'], color=color, linestyle='none', label=nomes_abas[i]) # lines armazenam as linhas geradas pelos plots
    line2, = plt.plot(x2, y, color='black', linestyle='-', linewidth=1)

    lines.append(line1)

plt.semilogy()
plt.ticklabel_format(axis='x', style='sci', scilimits=(0,0)) # eixo x em notação científica
plt.legend(handles=lines, loc='upper right', fontsize=4, markerscale=0.5) # cria a legenda usando as linhas armazenadas na lista
plt.xlabel('1/T [K$^{-1}$]',fontproperties=font)
plt.ylabel('ln P [Mpa]',fontproperties=font)
#plt.legend(loc="upper left")

nome_arquivo_python = os.path.basename(__file__) # obter o nome do arquivo python atual
nome_arquivo_sem_extensao = os.path.splitext(nome_arquivo_python)[0] # obter o nome do arquivo sem extensão # [0] vai acessar o primeiro elemento da tupla retornado por path.splitext

plt.title(nome_arquivo_sem_extensao)

plt.savefig(f'{nome_arquivo_sem_extensao}.png',format='png',bbox_inches='tight',
                dpi=300,transparent=False)
plt.show()