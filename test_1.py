#Plot_consistency
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.font_manager import FontProperties
import pandas as pd
import numpy as np
import openpyxl

matplotlib.rcParams['font.family'] = "Arial" # change the default font # mudança da fonte 
matplotlib.rcParams['xtick.direction'] = 'in' # change the ticks direction # mudança da direção dos marcadores ao longo dos eixos de um gráfico que indicam valores específicos 
matplotlib.rcParams['ytick.direction'] = 'in'
font = FontProperties() # cria uma instância de propriedades de fontes, tamanho, estilo e família
font.set_family('sans-serif') # 'serif', 'sans-serif', 'cursive', 'fantasy', or 'monospace' # família da fonte 
font.set_name('Arial') # nome da fonte 
font.set_weight('bold') # 'ultralight', 'light', 'normal', 'regular', 'book', 'medium', 'roman', 'semibold', 'demibold', 'demi', 'bold', 'heavy', 'extra bold', 'black' # peso da fonte 
font.set_style('normal') #'normal', 'italic' or 'oblique' # estilo da fonte, negrito, itálico, normal, etc 
font.set_size('x-large') # xx-small', 'x-small', 'small', 'medium', 'large', 'x-large', 'xx-large' # tamanho da fonte 
plt.rcParams['figure.dpi'] = 300 # resolução padrão para as figuras exibidas na tela 
plt.rcParams['savefig.dpi'] = 300 # resolução padrão para as figuras salvas em arquivos PNG, PDF ou SVG 
opt = {'fillstyle':'full','markersize':5,'markeredgewidth':0.5} # aparência dos marcadores, fillstyle: full - estilo de preenchimento markersize: 5 - tamanho do marcador markeredgewidth: 0.5 largura da borda do marcador 
plt.rcParams['figure.facecolor'] = (1.0,1.0,1.0,0.0) # cor de fundo da figura como transparente 
plt.rcParams['axes.facecolor'] = '0.95' # cor de fundo dos eixos como uma tonalidade de cinza claro 
#plt.rcParams['figure.figsize'] = (4,4) # tamanho da figura padrão em polegadas 


caminho_arquivo_dma = 'DMA.xlsx' # caminho do arquivo de DMA 
wb = openpyxl.load_workbook(caminho_arquivo_dma) # ler dados de abas individuais do arquivo Excel sem abrir e fechar repetidamente, mantém aberto. Mais eficiente se você precisa ler dados de várias abas do arquivo. Lista com os nomes de todas as abas do arquivo Excel.
# load_workbook é a função usada para carregar o arquvio Excel

dados_coluna_C_array = [] # matriz para guardar os dados da coluna C de todas as sheet
dados_coluna_I_array = [] # matriz para guardar os dados da coluna I de todas as sheet

for i in wb.sheetnames: # para i dentro da lista wb com os nomes de todas as abas
    nome_aba = wb[i] # obtém o nome da aba na posição i na lista wb, que contém os nomes de todas as abas

    dados_coluna_C = [] # abre uma lista para os dados da coluna C da sheet 
    dados_coluna_I = [] # abre uma lista para os dados da coluna I da sheet

    for j in nome_aba.iter_rows(min_row=2, values_only=True): # iter_rows: permite iterar sobre as linhas dessa aba # min_row=2: especifica a primeira linha a ser considerada, nesse caso, começa na segunda linha, pois a primeira linha geralmente contém os cabeçalhos das colunas # values_only=True: indica que queremos obter apenas os valores das células, sem incluir informações adicionais como formatação, fórmulas ou estilos. 
        dados_coluna_C.append(j[2])
        dados_coluna_I.append(j[8])

    dados_coluna_C_array.append(dados_coluna_C)
    dados_coluna_I_array.append(dados_coluna_I)

print('dados_C', dados_coluna_C_array)
print('dados_I', dados_coluna_I_array)

# Plotagem dos dados:

symbols = ['o', 's', '^', 'D', 'v', '>', '<', 'P', 'X', 'H']  # Exemplos de símbolos

# Vetor de cores
colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']  # Exemplos de cores

for i in range(len(dados_coluna_C_array)):
    x = dados_coluna_C_array[i]
    y = dados_coluna_I_array[i]
    print('x', x)
    print('y', y)
    for j in range(len(symbols)):
        symbol=symbols[i]
        color=colors[i]
        plt.plot(x, y, marker=symbol, markersize=opt['markersize'], markeredgewidth=opt['markeredgewidth'], fillstyle=opt['fillstyle'], color=color, linestyle='none')


plt.xlabel('Temperature [K]',fontproperties=font)
plt.ylabel('$\Delta$T/(T$_0$T)',fontproperties=font)
#plt.legend(loc="upper left")

plt.savefig('test1_EMIM_10.png',format='png',bbox_inches='tight',
                dpi=300,transparent=False)
plt.show()    